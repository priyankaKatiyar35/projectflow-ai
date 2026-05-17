"""
app/routes/reports.py
CMMI-compliant audit reports. Three formats: CSV, Excel (.xlsx), PDF.

Filters supported (all query params):
  ?date_from=YYYY-MM-DD
  ?date_to=YYYY-MM-DD
  ?employee_id=N
  ?project_id=N
  ?status=not_started|in_progress|completed|on_hold

Permissions:
  Admin    -> sees all data
  Employee -> only their assigned tasks (employee_id is forced to themselves)
"""
import io
import csv
from datetime import date, datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import ProjectTask, Project, User
from app.routes.auth import current_user


router = APIRouter(prefix="/api/reports", tags=["reports"])


# ---------- Column definitions (single source of truth) ----------

COLUMNS = [
    ("Project",         lambda t: t.project.name if t.project else ""),
    ("Phase",           lambda t: t.milestone_phase or ""),
    ("Category",        lambda t: t.category or ""),
    ("Action Item",     lambda t: t.action_item or ""),
    ("Task Description",lambda t: t.task_description or ""),
    ("Assigned To",     lambda t: ", ".join(a.full_name for a in t.assignees)),
    ("Planned Start",   lambda t: t.planned_start_date.isoformat() if t.planned_start_date else ""),
    ("Planned End",     lambda t: t.planned_end_date.isoformat()   if t.planned_end_date   else ""),
    ("Actual Start",    lambda t: t.actual_start_date.isoformat()  if t.actual_start_date  else ""),
    ("Actual End",      lambda t: t.actual_end_date.isoformat()    if t.actual_end_date    else ""),
    ("Planned Effort (h)", lambda t: t.planned_effort or 0),
    ("Actual Effort (h)",  lambda t: t.actual_effort  or 0),
    ("Variance %",      lambda t: t.effort_variance_pct),
    ("Days Delayed",    lambda t: t.days_delayed),
    ("Status",          lambda t: (t.status or "").replace("_", " ").title()),
    ("Remarks",         lambda t: t.remarks or ""),
    ("Created",         lambda t: t.created_at.strftime("%Y-%m-%d") if t.created_at else ""),
]


def _fetch_tasks(
    db: Session,
    user: User,
    date_from: Optional[date],
    date_to: Optional[date],
    employee_id: Optional[int],
    project_id: Optional[int],
    status: Optional[str],
):
    q = db.query(ProjectTask).join(Project)

    # Employees can only see their own
    if user.role != "admin":
        q = q.join(ProjectTask.assignees).filter(User.id == user.id)
    elif employee_id:
        q = q.join(ProjectTask.assignees).filter(User.id == employee_id)

    if project_id:
        q = q.filter(ProjectTask.project_id == project_id)
    if status:
        q = q.filter(ProjectTask.status == status)
    if date_from:
        q = q.filter(ProjectTask.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.filter(ProjectTask.created_at <= datetime.combine(date_to, datetime.max.time()))

    return q.order_by(ProjectTask.created_at.desc()).distinct().all()


# ============ JSON preview (for the report page UI) ============

@router.get("/cmmi/preview")
def preview_report(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    employee_id: Optional[int] = None,
    project_id:  Optional[int] = None,
    status:      Optional[str] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    tasks = _fetch_tasks(db, user, date_from, date_to, employee_id, project_id, status)
    rows = [{col: fn(t) for col, fn in COLUMNS} for t in tasks]

    # Summary stats for the page header
    total = len(rows)
    completed = sum(1 for t in tasks if t.status == "completed")
    delayed = sum(1 for t in tasks if t.days_delayed > 0)
    avg_variance = round(sum(t.effort_variance_pct for t in tasks) / total, 1) if total else 0

    return {
        "rows": rows,
        "columns": [c for c, _ in COLUMNS],
        "summary": {
            "total": total,
            "completed": completed,
            "delayed": delayed,
            "avg_variance_pct": avg_variance,
        },
        "generated_at": datetime.utcnow().isoformat(),
        "filters_applied": {
            "date_from": date_from.isoformat() if date_from else None,
            "date_to":   date_to.isoformat()   if date_to else None,
            "employee_id": employee_id,
            "project_id":  project_id,
            "status":      status,
        },
    }


# ============ CSV export ============

@router.get("/cmmi/csv")
def export_csv(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    employee_id: Optional[int] = None,
    project_id:  Optional[int] = None,
    status:      Optional[str] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    tasks = _fetch_tasks(db, user, date_from, date_to, employee_id, project_id, status)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c for c, _ in COLUMNS])
    for t in tasks:
        writer.writerow([fn(t) for _, fn in COLUMNS])

    buf.seek(0)
    fname = f"cmmi-action-items-{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ============ Excel export ============

@router.get("/cmmi/xlsx")
def export_xlsx(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    employee_id: Optional[int] = None,
    project_id:  Optional[int] = None,
    status:      Optional[str] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    tasks = _fetch_tasks(db, user, date_from, date_to, employee_id, project_id, status)

    wb = Workbook()
    ws = wb.active
    ws.title = "CMMI Action Items"

    # Title row
    ws.cell(row=1, column=1, value="CMMI Audit — Action Items Tracker")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="667EEA")
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(COLUMNS))
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata row
    ws.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  |  Rows: {len(tasks)}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(COLUMNS))

    # Header row
    header_row = 4
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for i, (col, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=header_row, column=i, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # Data rows
    for r, t in enumerate(tasks, start=header_row + 1):
        for i, (_, fn) in enumerate(COLUMNS, 1):
            value = fn(t)
            c = ws.cell(row=r, column=i, value=value)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

            # Conditional formatting: red if delayed or variance > 20%
            col_name = COLUMNS[i-1][0]
            if col_name == "Days Delayed" and isinstance(value, (int, float)) and value > 0:
                c.fill = PatternFill("solid", fgColor="FECACA")
                c.font = Font(color="991B1B", bold=True)
            elif col_name == "Variance %" and isinstance(value, (int, float)) and abs(value) > 20:
                c.fill = PatternFill("solid", fgColor="FEE2E2")
                c.font = Font(color="991B1B", bold=True)
            elif col_name == "Status":
                if value == "Completed":
                    c.fill = PatternFill("solid", fgColor="D1FAE5")
                    c.font = Font(color="065F46")
                elif value == "In Progress":
                    c.fill = PatternFill("solid", fgColor="FEF3C7")
                    c.font = Font(color="92400E")

    # Column widths
    widths = [20, 14, 14, 28, 36, 22, 12, 12, 12, 12, 13, 13, 11, 12, 13, 28, 12]
    for i, w in enumerate(widths[:len(COLUMNS)], 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"cmmi-action-items-{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ============ PDF export ============

@router.get("/cmmi/pdf")
def export_pdf(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    employee_id: Optional[int] = None,
    project_id:  Optional[int] = None,
    status:      Optional[str] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    try:
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")

    tasks = _fetch_tasks(db, user, date_from, date_to, employee_id, project_id, status)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#4F46E5"))
    meta_style  = ParagraphStyle("meta",  parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    cell_style  = ParagraphStyle("cell",  parent=styles["Normal"], fontSize=7, leading=9)

    elements = []
    elements.append(Paragraph("CMMI Audit — Action Items Tracker", title_style))
    meta = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  |  Rows: {len(tasks)}"
    if user.role != "admin":
        meta += f"  |  Employee: {user.full_name}"
    elements.append(Paragraph(meta, meta_style))
    elements.append(Spacer(1, 8))

    # Use Paragraphs so long text wraps inside cells
    data = [[Paragraph(c, cell_style) for c, _ in COLUMNS]]
    for t in tasks:
        row = []
        for col, fn in COLUMNS:
            v = fn(t)
            row.append(Paragraph(str(v) if v not in (None, "") else "—", cell_style))
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 7),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    doc.build(elements)

    pdf_bytes = buf.getvalue()
    buf.close()
    fname = f"cmmi-action-items-{date.today().isoformat()}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
